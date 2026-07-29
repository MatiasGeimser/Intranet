"""Crea usuarios del archivo FTP y asigna carpetas privadas por persona."""

import os
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.database import SessionLocal
from app.core.security import get_password_hash
from app.models.folder_access import FolderAccess
from app.models.role import Role
from app.models.user import User

EMAIL_DOMAIN = "natura.cl"
LEGACY_EMAIL_DOMAIN = "geoinfobusinsess.cl"
INITIAL_PASSWORD = os.environ.get("NATURA_INITIAL_PASSWORD")
SKIPPED_SECTIONS = {"NATURA", "FACTURACION", "REPORTES"}


def clean_name(value: str) -> str:
    name = " ".join(str(value).strip().split())
    replacements = {"NU�EZ": "NUÑEZ", "MAR�A": "MARÍA"}
    for original, corrected in replacements.items():
        name = name.replace(original, corrected)
    return name.title()


def make_email(full_name: str) -> str:
    normalized = unicodedata.normalize("NFKD", full_name).encode("ascii", "ignore").decode("ascii")
    tokens = re.findall(r"[a-z]+", normalized.lower())
    return f"{tokens[0]}.{tokens[-1]}@{EMAIL_DOMAIN}"


def make_legacy_email(full_name: str) -> str:
    normalized = unicodedata.normalize("NFKD", full_name).encode("ascii", "ignore").decode("ascii")
    tokens = re.findall(r"[a-z]+", normalized.lower())
    return f"{tokens[0]}.{tokens[-1]}@{LEGACY_EMAIL_DOMAIN}"


def read_people(workbook_path: Path) -> dict[str, list[str]]:
    sheet = load_workbook(workbook_path, read_only=True, data_only=True).active
    people: dict[str, list[str]] = {}
    current_name: str | None = None

    for _, name_cell, folder_cell in sheet.iter_rows(values_only=True):
        if name_cell:
            candidate = clean_name(name_cell)
            if candidate.upper() in SKIPPED_SECTIONS:
                current_name = None
                continue
            if folder_cell:
                current_name = candidate
                people[current_name] = [str(folder_cell).strip()]
                continue
        elif current_name and folder_cell:
            people[current_name].append(str(folder_cell).strip())

    return people


def main() -> None:
    workbook_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(r"C:\Users\User\Downloads\FTP.xlsx")
    people = read_people(workbook_path)
    if not people:
        raise RuntimeError("No se encontraron personas en el Excel.")
    if not INITIAL_PASSWORD:
        raise RuntimeError("Define NATURA_INITIAL_PASSWORD antes de ejecutar la importación.")

    db = SessionLocal()
    created = updated = 0
    try:
        user_role = db.query(Role).filter(Role.name == "Usuario").one()
        for full_name, subfolders in people.items():
            email = make_email(full_name)
            user = (
                db.query(User)
                .filter(func.lower(User.email).in_([email.lower(), make_legacy_email(full_name).lower()]))
                .first()
            )
            if user is None:
                user = User(
                    email=email,
                    full_name=full_name,
                    hashed_password=get_password_hash(INITIAL_PASSWORD),
                    role_id=user_role.id,
                    is_active=True,
                    gender="Mujer",
                    avatar_url="/static/uploads/avatars/woman.png",
                )
                db.add(user)
                db.flush()
                created += 1
            else:
                user.email = email
                user.full_name = full_name
                user.role_id = user_role.id
                user.is_active = True
                updated += 1

            # La cuenta conserva exclusivamente las carpetas personales indicadas por el Excel.
            db.query(FolderAccess).filter(FolderAccess.user_id == user.id).delete()
            for subfolder in dict.fromkeys(subfolders):
                db.add(FolderAccess(
                    user_id=user.id,
                    folder_name=f"Natura / CBE / {full_name} / {subfolder}",
                    can_read=True,
                    can_write=False,
                ))

        db.commit()
        print(f"created={created} updated={updated} people={len(people)}")
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
