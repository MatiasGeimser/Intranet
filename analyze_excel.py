import openpyxl

wb = openpyxl.load_workbook(r'static\uploads\General\Detalle Red Geimser (1).xlsx')
ws = wb.active

output = []
output.append(f"SHEET: {ws.title}\n")
output.append(f"ROWS: {ws.max_row}, COLS: {ws.max_column}\n")
output.append("="*150 + "\n\n")

for row_idx in range(1, ws.max_row + 1):
    row_vals = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row_idx, col_idx).value
        row_vals.append(str(val) if val is not None else "")
    output.append(f"R{row_idx:02d}: {' | '.join(row_vals)}\n")

with open('excel_analysis.txt', 'w', encoding='utf-8') as f:
    f.writelines(output)

print("Análisis guardado en excel_analysis.txt")
