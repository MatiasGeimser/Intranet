import re

file_path = 'c:/Intranet/app/api/endpoints/duplicate_phones.py'
with open(file_path, 'r', encoding='utf-8') as f:
    content = f.read()

new_function = '''@router.post("/process")
async def process_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    file2: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")
        
    if file2 and not file2.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="La base secundaria debe ser un Excel (.xlsx)")

    seen_phones = set()
    total_deleted = 0
    temp_input_paths = []
    
    import re
    import uuid
    import os
    import tempfile
    import openpyxl

    try:
        # Paso 1: Si hay Base 2, extraer todos los telefonos y guardarlos en seen_phones (blacklist)
        if file2:
            content2 = await file2.read()
            temp_input_path2 = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
            temp_input_paths.append(temp_input_path2)
            with open(temp_input_path2, "wb") as f:
                f.write(content2)
                
            wb2 = openpyxl.load_workbook(temp_input_path2, read_only=True)
            for sheetname in wb2.sheetnames:
                sheet = wb2[sheetname]
                for row in sheet.iter_rows(values_only=True):
                    for val in row:
                        if val is not None:
                            val_str = str(val).strip()
                            parts = re.split(r'[/;,]|\\s+y\\s+|\\s+-\\s+', val_str)
                            for part in parts:
                                cleaned = ''.join(filter(str.isdigit, part))
                                if 7 <= len(cleaned) <= 15 and part.replace('+','').replace('-','').replace(' ','').strip().isdigit():
                                    seen_phones.add(cleaned)
            wb2.close()
            remove_temp_file(temp_input_path2)
            temp_input_paths.remove(temp_input_path2)

        # Paso 2: Procesar la Base 1 normal y limpiarla de sus duplicados + los de Base 2
        content1 = await file.read()
        temp_input_path = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
        temp_input_paths.append(temp_input_path)
        with open(temp_input_path, "wb") as f:
            f.write(content1)

        wb = openpyxl.load_workbook(temp_input_path, read_only=True)
        new_wb = openpyxl.Workbook(write_only=True)
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            new_sheet = new_wb.create_sheet(title=sheet.title)
            
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
                for val in header:
                    if val is not None:
                        val_str = str(val).strip()
                        cleaned_val = ''.join(filter(str.isdigit, val_str))
                        if 7 <= len(cleaned_val) <= 15 and val_str.replace('+','').replace('-','').replace(' ','').isdigit():
                            seen_phones.add(cleaned_val)
            except StopIteration:
                continue
                
            processed_rows = []
            max_extra_phones = 0
            
            for row in rows_iter:
                row_has_phone = False
                is_duplicate = False
                phones_in_row = []
                
                new_row = list(row)
                extra_phones_for_this_row = []
                
                for col_idx, val in enumerate(row):
                    if val is not None:
                        val_str = str(val).strip()
                        parts = re.split(r'[/;,]|\\s+y\\s+|\\s+-\\s+', val_str)
                        
                        valid_parts = []
                        for part in parts:
                            cleaned = ''.join(filter(str.isdigit, part))
                            if 7 <= len(cleaned) <= 15 and part.replace('+','').replace('-','').replace(' ','').strip().isdigit():
                                valid_parts.append((part.strip(), cleaned))
                                
                        if len(valid_parts) > 0:
                            row_has_phone = True
                            new_row[col_idx] = valid_parts[0][0]
                            phones_in_row.append(valid_parts[0][1])
                            
                            for extra_part, extra_cleaned in valid_parts[1:]:
                                extra_phones_for_this_row.append(extra_part)
                                phones_in_row.append(extra_cleaned)
                                
                for p in phones_in_row:
                    if p in seen_phones:
                        is_duplicate = True
                        break
                
                if row_has_phone and is_duplicate:
                    total_deleted += 1
                else:
                    for p in phones_in_row:
                        seen_phones.add(p)
                    new_row.extend(extra_phones_for_this_row)
                    max_extra_phones = max(max_extra_phones, len(extra_phones_for_this_row))
                    processed_rows.append(new_row)
                    
            for i in range(max_extra_phones):
                header.append(f"Teléfono {i+2}")
                
            new_sheet.append(header)
            for pr in processed_rows:
                new_sheet.append(pr)
        
        wb.close()
        temp_output_path = os.path.join(TEMP_DIR, f"clean_pri_{uuid.uuid4().hex}.xlsx")
        new_wb.save(temp_output_path)
        
        for path in temp_input_paths:
            remove_temp_file(path)
            
        msg = "Archivo procesado exitosamente." if not file2 else "Cruce exitoso: Base 1 ha sido limpiada usando Base 2."
        download_id = os.path.basename(temp_output_path)
        original_filename = file.filename
        
        from fastapi.responses import JSONResponse
        return JSONResponse({
            "message": msg,
            "deleted_count": total_deleted,
            "download_id": download_id,
            "original_filename": original_filename
        })
        
    except Exception as e:
        for path in temp_input_paths:
            remove_temp_file(path)
        print(f"Error procesando el Excel: {e}")
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(e)}")
'''

# Replace between @router.post("/process") and @router.post("/process-commercial")
pattern = re.compile(r'@router\.post\("/process"\).*?(?=@router\.post\("/process-commercial"\))', re.DOTALL)
new_content = pattern.sub(new_function + "\n", content)

with open(file_path, 'w', encoding='utf-8') as f:
    f.write(new_content)
