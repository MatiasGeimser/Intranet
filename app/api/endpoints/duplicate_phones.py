from fastapi import APIRouter, Depends, UploadFile, File, Form, HTTPException, Request, BackgroundTasks
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.api.deps import get_current_active_user
from app.models.user import User
import openpyxl
import os
import uuid
import tempfile
import zipfile
from typing import Optional

router = APIRouter()

# Directorio temporal para los archivos generados
TEMP_DIR = os.path.join(tempfile.gettempdir(), "intranet_excel")
os.makedirs(TEMP_DIR, exist_ok=True)

def remove_temp_file(path: str):
    try:
        if os.path.exists(path):
            os.remove(path)
    except Exception as e:
        print(f"No se pudo eliminar el archivo temporal {path}: {e}")

@router.post("/process")
async def process_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    file2: Optional[UploadFile] = File(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")
        
    if file2 and not file2.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="La base secundaria debe ser un Excel (.xlsx)")

    seen_phones = set()
    total_deleted = 0
    temp_input_paths = []
    
    import re

    try:
        # Paso 1: Si hay Base 2, extraer todos los telefonos y guardarlos en seen_phones (blacklist)
        if file2:
            content2 = await file2.read()
            temp_input_path2 = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
            temp_input_paths.append(temp_input_path2)
            with open(temp_input_path2, "wb") as f:
                f.write(content2)
                
            wb2 = openpyxl.load_workbook(temp_input_path2, read_only=True)
            for sheetname in wb2.sheetnames:
                sheet = wb2[sheetname]
                for row in sheet.iter_rows(values_only=True):
                    for val in row:
                        if val is not None:
                            val_str = str(val).strip()
                            parts = re.split(r'[/;,]|\s+y\s+|\s+-\s+', val_str)
                            for part in parts:
                                cleaned = ''.join(filter(str.isdigit, part))
                                if 7 <= len(cleaned) <= 15 and part.replace('+','').replace('-','').replace(' ','').strip().isdigit():
                                    seen_phones.add(cleaned)
            wb2.close()
            remove_temp_file(temp_input_path2)
            temp_input_paths.remove(temp_input_path2)

        # Paso 2: Procesar la Base 1 normal y limpiarla de sus duplicados + los de Base 2
        content1 = await file.read()
        temp_input_path = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
        temp_input_paths.append(temp_input_path)
        with open(temp_input_path, "wb") as f:
            f.write(content1)

        wb = openpyxl.load_workbook(temp_input_path, read_only=True)
        new_wb = openpyxl.Workbook(write_only=True)
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            new_sheet = new_wb.create_sheet(title=sheet.title)
            
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
                for val in header:
                    if val is not None:
                        val_str = str(val).strip()
                        cleaned_val = ''.join(filter(str.isdigit, val_str))
                        if 7 <= len(cleaned_val) <= 15 and val_str.replace('+','').replace('-','').replace(' ','').isdigit():
                            seen_phones.add(cleaned_val)
            except StopIteration:
                continue
                
            processed_rows = []
            max_extra_phones = 0
            
            for row in rows_iter:
                row_has_phone = False
                is_duplicate = False
                phones_in_row = []
                
                new_row = list(row)
                extra_phones_for_this_row = []
                
                for col_idx, val in enumerate(row):
                    if val is not None:
                        val_str = str(val).strip()
                        parts = re.split(r'[/;,]|\s+y\s+|\s+-\s+', val_str)
                        
                        valid_parts = []
                        for part in parts:
                            cleaned = ''.join(filter(str.isdigit, part))
                            if 7 <= len(cleaned) <= 15 and part.replace('+','').replace('-','').replace(' ','').isdigit():
                                valid_parts.append((part.strip(), cleaned))
                                
                        if len(valid_parts) > 0:
                            row_has_phone = True
                            # Conservar el primero en su lugar original
                            new_row[col_idx] = valid_parts[0][0]
                            phones_in_row.append(valid_parts[0][1])
                            
                            # Los adicionales se extraen para nuevas columnas
                            for extra_part, extra_cleaned in valid_parts[1:]:
                                extra_phones_for_this_row.append(extra_part)
                                phones_in_row.append(extra_cleaned)
                                
                for p in phones_in_row:
                    if p in seen_phones:
                        is_duplicate = True
                        break
                
                if row_has_phone and is_duplicate:
                    total_deleted += 1
                else:
                    for p in phones_in_row:
                        seen_phones.add(p)
                    new_row.extend(extra_phones_for_this_row)
                    max_extra_phones = max(max_extra_phones, len(extra_phones_for_this_row))
                    processed_rows.append(new_row)
                    
            for i in range(max_extra_phones):
                header.append(f"Teléfono {i+2}")
                
            new_sheet.append(header)
            for pr in processed_rows:
                new_sheet.append(pr)
        
        wb.close()
        temp_output_path = os.path.join(TEMP_DIR, f"clean_pri_{uuid.uuid4().hex}.xlsx")
        new_wb.save(temp_output_path)
        
        for path in temp_input_paths:
            remove_temp_file(path)
            
        msg = "Archivo procesado exitosamente." if not file2 else "Cruce exitoso: Base 1 ha sido limpiada usando Base 2 como referencia."
        download_id = os.path.basename(temp_output_path)
        original_filename = file.filename
        
        return JSONResponse({
            "message": msg,
            "deleted_count": total_deleted,
            "download_id": download_id,
            "original_filename": original_filename
        })
        
    except Exception as e:
        for path in temp_input_paths:
            remove_temp_file(path)
        print(f"Error procesando el Excel: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(e)}")

@router.post("/process-commercial")
async def process_commercial_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    column_name: Optional[str] = Form(None),
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    target_cols = [c.lower().strip() for c in (column_name.split(",") if column_name else ["nombre comercial", "empresa", "razon social"])]
    
    seen_phones_by_company = set()
    total_deleted = 0
    temp_input_path = os.path.join(TEMP_DIR, f"input_{uuid.uuid4().hex}.xlsx")
    
    try:
        content = await file.read()
        with open(temp_input_path, "wb") as f:
            f.write(content)

        wb = openpyxl.load_workbook(temp_input_path, read_only=True)
        new_wb = openpyxl.Workbook(write_only=True)
        import re
        
        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            new_sheet = new_wb.create_sheet(title=sheet.title)
            
            rows_iter = sheet.iter_rows(values_only=True)
            try:
                header = list(next(rows_iter))
            except StopIteration:
                continue
            
            company_col_idx = -1
            if header:
                for idx, h in enumerate(header):
                    if h is not None:
                        h_str = str(h).lower().strip()
                        if any(tc in h_str for tc in target_cols if tc):
                            company_col_idx = idx
                            break
                            
            if company_col_idx == -1:
                col_names_str = column_name if column_name else "Nombre Comercial, Empresa, o Razon Social"
                raise HTTPException(status_code=400, detail=f"No se encontró una columna de nombre: '{col_names_str}' en la hoja {sheet.title}.")

            processed_rows = []
            max_extra_phones = 0
            
            for row in rows_iter:
                row_has_phone = False
                is_duplicate = False
                phones_in_row = []
                
                company_val = row[company_col_idx] if company_col_idx < len(row) else ""
                company_name = str(company_val).strip().lower() if company_val is not None else ""
                
                new_row = list(row)
                extra_phones_for_this_row = []
                
                for col_idx, val in enumerate(row):
                    if col_idx == company_col_idx:
                        continue
                        
                    if val is not None:
                        val_str = str(val).strip()
                        parts = re.split(r'[/;,]|\s+y\s+|\s+-\s+', val_str)
                        
                        valid_parts = []
                        for part in parts:
                            cleaned = ''.join(filter(str.isdigit, part))
                            if 7 <= len(cleaned) <= 15 and part.replace('+','').replace('-','').replace(' ','').strip().isdigit():
                                valid_parts.append((part.strip(), cleaned))
                                
                        if len(valid_parts) > 0:
                            row_has_phone = True
                            new_row[col_idx] = valid_parts[0][0]
                            phones_in_row.append(valid_parts[0][1])
                            
                            for extra_part, extra_cleaned in valid_parts[1:]:
                                extra_phones_for_this_row.append(extra_part)
                                phones_in_row.append(extra_cleaned)
                                
                for p in phones_in_row:
                    if (company_name, p) in seen_phones_by_company:
                        is_duplicate = True
                        break
                
                if row_has_phone and is_duplicate:
                    total_deleted += 1
                else:
                    for p in phones_in_row:
                        seen_phones_by_company.add((company_name, p))
                    new_row.extend(extra_phones_for_this_row)
                    max_extra_phones = max(max_extra_phones, len(extra_phones_for_this_row))
                    processed_rows.append(new_row)
                    
            for i in range(max_extra_phones):
                header.append(f"Teléfono {i+2}")
                
            new_sheet.append(header)
            for pr in processed_rows:
                new_sheet.append(pr)
        
        wb.close()
        temp_output_path = os.path.join(TEMP_DIR, f"clean_pri_{uuid.uuid4().hex}.xlsx")
        new_wb.save(temp_output_path)
        remove_temp_file(temp_input_path)
        
        return JSONResponse({
            "message": "Archivo procesado exitosamente.",
            "deleted_count": total_deleted,
            "download_id": os.path.basename(temp_output_path),
            "original_filename": file.filename
        })
        
    except HTTPException:
        remove_temp_file(temp_input_path)
        raise
    except Exception as e:
        remove_temp_file(temp_input_path)
        print(f"Error procesando el Excel comercial: {e}")
        raise HTTPException(status_code=500, detail=f"Error procesando el archivo: {str(e)}")


@router.get("/download/{download_id}")
async def download_excel(
    download_id: str,
    background_tasks: BackgroundTasks,
    filename: Optional[str] = None,
    current_user: User = Depends(get_current_active_user)
):
    # Validar que el archivo exista en nuestro TEMP_DIR para evitar path traversal
    file_path = os.path.join(TEMP_DIR, download_id)
    if not os.path.exists(file_path) or not (download_id.startswith("clean_") or download_id.startswith("clean_cross_")) or ".." in download_id:
        raise HTTPException(status_code=404, detail="Archivo no encontrado o expirado")
        
    background_tasks.add_task(remove_temp_file, file_path)
    
    if download_id.endswith(".zip"):
        media_type = "application/zip"
        download_name = "bases_cruzadas_limpias.zip"
    else:
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if filename:
            name, ext = os.path.splitext(filename)
            download_name = f"{name}_LIMPIO{ext}"
        else:
            download_name = "telefonos_limpios.xlsx"
    
    return FileResponse(
        path=file_path,
        filename=download_name,
        media_type=media_type
    )
