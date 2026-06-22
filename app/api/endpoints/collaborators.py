from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Request
from sqlalchemy.orm import Session
from sqlalchemy import or_, and_
from typing import List, Optional
from io import BytesIO
from openpyxl import Workbook, load_workbook

from app.core.database import get_db
from app.models.collaborator import Collaborator
from app.schemas.collaborator import CollaboratorCreate, CollaboratorUpdate, CollaboratorOut
from app.api.deps import get_current_active_user, PermissionChecker
from app.models.user import User

router = APIRouter()

@router.get("/", response_model=List[CollaboratorOut])
def read_collaborators(
    q: Optional[str] = None,
    company: Optional[str] = None,
    area: Optional[str] = None,
    department: Optional[str] = None,
    status_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """
    Get all collaborators with advanced search and filters.
    """
    query = db.query(Collaborator)

    if q:
        search_term = f"%{q}%"
        query = query.filter(
            or_(
                Collaborator.full_name.ilike(search_term),
                Collaborator.position.ilike(search_term),
                Collaborator.company.ilike(search_term),
                Collaborator.area.ilike(search_term),
                Collaborator.department.ilike(search_term),
                Collaborator.email.ilike(search_term),
                Collaborator.phone.ilike(search_term),
                Collaborator.extension_3cx.ilike(search_term),
                Collaborator.direct_boss.ilike(search_term),
                Collaborator.branch.ilike(search_term),
            )
        )

    if company:
        query = query.filter(Collaborator.company == company)
    if area:
        query = query.filter(Collaborator.area == area)
    if department:
        query = query.filter(Collaborator.department == department)
    if status_filter:
        query = query.filter(Collaborator.status == status_filter)

    return query.order_by(Collaborator.full_name.asc()).offset(skip).limit(limit).all()

@router.get("/{collab_id}", response_model=CollaboratorOut)
def read_collaborator(
    collab_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    collab = db.query(Collaborator).filter(Collaborator.id == collab_id).first()
    if not collab:
        raise HTTPException(status_code=404, detail="Colaborador no encontrado")
    return collab

@router.post("/", response_model=CollaboratorOut, status_code=status.HTTP_201_CREATED)
def create_collaborator(
    collab_in: CollaboratorCreate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("admin")) # Usually only admins or HR can manage this
):
    if current_user.role.name != "Administrador":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
        
    db_collab = Collaborator(**collab_in.dict())
    db.add(db_collab)
    db.commit()
    db.refresh(db_collab)
    return db_collab

@router.put("/{collab_id}", response_model=CollaboratorOut)
def update_collaborator(
    collab_id: int, 
    collab_in: CollaboratorUpdate, 
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("admin"))
):
    if current_user.role.name != "Administrador":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
        
    db_collab = db.query(Collaborator).filter(Collaborator.id == collab_id).first()
    if not db_collab:
        raise HTTPException(status_code=404, detail="Colaborador no encontrado")
    
    update_data = collab_in.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_collab, field, value)
        
    db.add(db_collab)
    db.commit()
    db.refresh(db_collab)
    return db_collab

@router.delete("/{collab_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_collaborator(
    collab_id: int, 
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("admin"))
):
    if current_user.role.name != "Administrador":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
        
    db_collab = db.query(Collaborator).filter(Collaborator.id == collab_id).first()
    if not db_collab:
        raise HTTPException(status_code=404, detail="Colaborador no encontrado")
        
    db.delete(db_collab)
    db.commit()
    return None

@router.get("/export")
def export_collaborators(
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("admin"))
):
    if current_user.role.name != "Administrador":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
        
    collaborators = db.query(Collaborator).all()
    
    data = []
    headers = ["Nombre", "Correo", "Cargo", "Empresa", "Área", "Departamento", "Jefe", "Teléfono", "Anexo", "Sucursal", "Dirección", "Estado", "Observaciones"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Directorio"
    ws.append(headers)
    
    for c in collaborators:
        row_data = [
            c.full_name, c.email, c.position, c.company, c.area, c.department,
            c.direct_boss, c.phone, c.extension_3cx, c.branch, c.address, c.status, c.observations
        ]
        ws.append(row_data)
        
    output = BytesIO()
    wb.save(output)
    
    from fastapi.responses import StreamingResponse
    headers = {
        'Content-Disposition': 'attachment; filename="directorio_corporativo.xlsx"'
    }
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

import os
from app.core.config import settings

@router.post("/{collab_id}/avatar")
def upload_collaborator_avatar(
    collab_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Permite al usuario cambiar la foto del colaborador si es admin o si el correo coincide."""
    db_collab = db.query(Collaborator).filter(Collaborator.id == collab_id).first()
    if not db_collab:
        raise HTTPException(status_code=404, detail="Colaborador no encontrado")
        
    if current_user.role.name != "Administrador":
        if not db_collab.email or db_collab.email.lower() != current_user.email.lower():
            raise HTTPException(status_code=403, detail="No tienes permisos para modificar este perfil.")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".png", ".jpg", ".jpeg", ".webp"]:
        raise HTTPException(status_code=400, detail="Formato de imagen inválido. Solo PNG, JPG, JPEG o WEBP.")

    avatar_dir = os.path.join(settings.UPLOAD_DIR, "avatars")
    os.makedirs(avatar_dir, exist_ok=True)

    filename = f"avatar_collab_{db_collab.id}{ext}"
    file_path = os.path.join(avatar_dir, filename)

    content = file.file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    relative_web_path = f"/static/uploads/avatars/{filename}"
    db_collab.avatar_url = relative_web_path
    
    # Optativo: si el correo coincide, también actualizar el avatar del User (CRM)
    if db_collab.email and db_collab.email.lower() == current_user.email.lower():
        current_user.avatar_url = relative_web_path
        
    db.commit()
    return {"avatar_url": relative_web_path}

@router.post("/import")
def import_collaborators(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("admin"))
):
    if current_user.role.name != "Administrador":
        raise HTTPException(status_code=403, detail="Permisos insuficientes")
        
    if not file.filename.endswith(('.xls', '.xlsx')):
        raise HTTPException(status_code=400, detail="Formato de archivo inválido. Se requiere Excel.")
        
    try:
        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        
        imported_count = 0
        updated_count = 0
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_data = dict(zip(headers, row))
            
            # Convert missing values to None
            for k, v in row_data.items():
                if v is None or str(v).strip() == "" or str(v).strip() == "nan":
                    row_data[k] = None
                    
            full_name = row_data.get("Nombre")
            if not full_name:
                continue
                
            email = row_data.get("Correo")
            
            existing = None
            if email:
                existing = db.query(Collaborator).filter(Collaborator.email == email).first()
            if not existing:
                existing = db.query(Collaborator).filter(Collaborator.full_name == full_name).first()
                
            if existing:
                existing.position = row_data.get("Cargo", existing.position)
                existing.company = row_data.get("Empresa", existing.company)
                existing.area = row_data.get("Área", existing.area)
                existing.department = row_data.get("Departamento", existing.department)
                existing.direct_boss = row_data.get("Jefe", existing.direct_boss)
                existing.email = email or existing.email
                existing.phone = str(row_data.get("Teléfono")) if row_data.get("Teléfono") else existing.phone
                existing.extension_3cx = str(row_data.get("Anexo")) if row_data.get("Anexo") else existing.extension_3cx
                existing.branch = row_data.get("Sucursal", existing.branch)
                existing.address = row_data.get("Dirección", existing.address)
                existing.status = row_data.get("Estado", existing.status) or "Disponible"
                updated_count += 1
            else:
                new_collab = Collaborator(
                    full_name=full_name,
                    position=row_data.get("Cargo"),
                    company=row_data.get("Empresa"),
                    area=row_data.get("Área"),
                    department=row_data.get("Departamento"),
                    direct_boss=row_data.get("Jefe"),
                    email=email,
                    phone=str(row_data.get("Teléfono")) if row_data.get("Teléfono") else None,
                    extension_3cx=str(row_data.get("Anexo")) if row_data.get("Anexo") else None,
                    branch=row_data.get("Sucursal"),
                    address=row_data.get("Dirección"),
                    status=row_data.get("Estado", "Disponible")
                )
                db.add(new_collab)
                imported_count += 1
                
        db.commit()
        return {"detail": f"Importación finalizada. Nuevos: {imported_count}, Actualizados: {updated_count}"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error al procesar el archivo: {str(e)}")
