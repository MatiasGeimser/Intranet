from fastapi import APIRouter, Depends, HTTPException, Request, Response, status, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy import or_
from typing import List, Optional
from app.core.database import get_db
from app.models.credential import Credential
from app.schemas.credential import CredentialResponse, CredentialCreate, CredentialUpdate, ExecutiveCredentialCreate
from app.api.deps import get_current_active_user, PermissionChecker
from app.services.crypto_service import crypto_service
from app.services.audit_service import audit_service
from app.models.user import User

router = APIRouter()

TECHNOLOGY_AREAS = {"Tecnología", "Tecnologia", "Tecnología (IT)", "Tecnologia (IT)", "IT"}
ADMINISTRATION_AREAS = {"Administración", "Administracion"}
INFRASTRUCTURE_CREDENTIAL_TERMS = ("wifi", "wi-fi", "cpanel", "c-panel")
EXECUTIVE_CATEGORIES = {"Correo Corporativo", "CRM", "Telefonía", "Sistemas"}


def is_technology_administrator(user: User) -> bool:
    return bool(
        user.role
        and user.role.name == "Administrador"
        and user.area
        and user.area.name in TECHNOLOGY_AREAS
    )


def can_manage_vault_folders(user: User) -> bool:
    """Las categorías visuales de la bóveda están reservadas para Administración."""
    return bool(
        user.role
        and user.role.name == "Administrador"
        and user.area
        and user.area.name in ADMINISTRATION_AREAS
    )


def vault_category_for(user: User, requested_category: str) -> str:
    """Evita que usuarios no autorizados creen o mantengan carpetas personales."""
    if requested_category in EXECUTIVE_CATEGORIES or can_manage_vault_folders(user):
        return requested_category
    return "General"


def is_infrastructure_credential(credential: Credential) -> bool:
    searchable = " ".join(filter(None, [credential.title, credential.url, credential.category])).casefold()
    return any(term in searchable for term in INFRASTRUCTURE_CREDENTIAL_TERMS)


def is_executive_credential(credential: Credential) -> bool:
    return credential.category in EXECUTIVE_CATEGORIES


def executive_credential_filter():
    return Credential.category.in_(EXECUTIVE_CATEGORIES)


def credential_is_visible_to(credential: Credential, user: User) -> bool:
    if is_infrastructure_credential(credential):
        return is_technology_administrator(user)
    if is_executive_credential(credential):
        return user.role.name in {"Administrador", "Supervisor"}
    return credential.owner_id == user.id


def credential_visibility_filter(user: User):
    owner_filter = Credential.owner_id == user.id
    if not is_technology_administrator(user):
        return owner_filter
    infrastructure_filter = or_(*[
        or_(
            Credential.title.ilike(f"%{term}%"),
            Credential.url.ilike(f"%{term}%"),
            Credential.category.ilike(f"%{term}%"),
        )
        for term in INFRASTRUCTURE_CREDENTIAL_TERMS
    ])
    return or_(owner_filter, infrastructure_filter)

@router.get("", response_model=List[CredentialResponse])
def get_credentials(
    category: Optional[str] = None,
    scope: str = "personal",
    include_inactive: bool = False,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Obtiene credenciales personales o corporativas por ejecutivo."""
    if scope not in {"personal", "executive"}:
        raise HTTPException(status_code=400, detail="Ámbito de credenciales no válido.")
    query = db.query(Credential)
    if not include_inactive:
        query = query.filter(Credential.is_active.is_(True))
    
    if scope == "executive":
        if current_user.role.name not in {"Administrador", "Supervisor"}:
            raise HTTPException(status_code=403, detail="No tienes acceso a las credenciales por ejecutivo.")
        query = query.filter(executive_credential_filter())
    else:
        query = query.filter(~executive_credential_filter())
        query = query.filter(credential_visibility_filter(current_user))
        
    if category:
        query = query.filter(Credential.category == category)
        
    return query.order_by(Credential.created_at.desc()).all()


@router.post("", response_model=CredentialResponse, status_code=status.HTTP_201_CREATED)
def create_credential(
    request: Request,
    cred_data: CredentialCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Crea y guarda una credencial encriptando la contraseña con AES-256-GCM."""
    encrypted_pw = crypto_service.encrypt_password(cred_data.password)
    
    db_cred = Credential(
        title=cred_data.title,
        url=cred_data.url,
        username=cred_data.username,
        encrypted_password=encrypted_pw,
        category=vault_category_for(current_user, cred_data.category),
        owner_id=current_user.id
    )
    db.add(db_cred)
    db.commit()
    db.refresh(db_cred)

    # Auditoría
    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="credential_create",
        ip_address=request.client.host if request.client else None,
        details=f"Creó una nueva credencial en la bóveda: {db_cred.title} (Categoría: {db_cred.category})"
    )

    return db_cred


@router.get("/{cred_id}/decrypt")
def decrypt_credential_password(
    cred_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Desencripta de forma segura una credencial y registra obligatoriamente la auditoría de acceso."""
    cred = db.query(Credential).filter(Credential.id == cred_id).first()
    if not cred:
        raise HTTPException(status_code=404, detail="Credencial no encontrada.")
    if cred.is_active is False and current_user.role.name != "Administrador":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Credencial inactiva. No se permite revelar la contraseña."
        )
        
    if not credential_is_visible_to(cred, current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Acceso denegado. No tienes permisos para ver esta credencial."
        )

    # Desencriptar e inyectar auditoría
    decrypted_pw = crypto_service.decrypt_password(
        encrypted_password=cred.encrypted_password,
        db=db,
        user_id=current_user.id,
        credential_id=cred.id,
        ip_address=request.client.host if request.client else None
    )

    return {"password": decrypted_pw}


@router.put("/{cred_id}", response_model=CredentialResponse)
def update_credential(
    cred_id: int,
    request: Request,
    cred_data: CredentialUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Actualiza los datos de una credencial en la bóveda."""
    cred = db.query(Credential).filter(Credential.id == cred_id).first()
    if not cred:
        raise HTTPException(status_code=404, detail="Credencial no encontrada.")
        
    if not credential_is_visible_to(cred, current_user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado.")

    if cred_data.title:
        cred.title = cred_data.title
    if cred_data.url is not None:
        cred.url = cred_data.url
    if cred_data.username:
        cred.username = cred_data.username
    if cred_data.category:
        cred.category = vault_category_for(current_user, cred_data.category)
    if cred_data.is_active is not None:
        cred.is_active = cred_data.is_active
    if cred_data.password:
        cred.encrypted_password = crypto_service.encrypt_password(cred_data.password)

    db.commit()
    db.refresh(cred)

    # Auditoría
    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="credential_update",
        ip_address=request.client.host if request.client else None,
        details=f"Modificó la credencial ID {cred.id} ({cred.title})."
    )

    return cred


from pydantic import BaseModel
class ExecutiveStatusUpdate(BaseModel):
    is_active: bool

@router.put("/executive/{person_name}/status")
def update_executive_status(
    person_name: str,
    status_update: ExecutiveStatusUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("credentials:manage"))
):
    """Actualiza el estado de activo/inactivo para todas las credenciales de un ejecutivo."""
    query = db.query(Credential).filter(Credential.owner_id == current_user.id)
        
    from sqlalchemy import or_
    creds = query.filter(or_(
        Credential.title.endswith(f" - {person_name}"),
        Credential.title.like(f"% - {person_name} (%)"),
        Credential.title == person_name
    )).all()
    
    if not creds:
        raise HTTPException(status_code=404, detail="Ejecutivo no encontrado en la bóveda.")
        
    for c in creds:
        c.is_active = status_update.is_active
        
    db.commit()
    
    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="executive_status_update",
        ip_address=request.client.host if request.client else None,
        details=f"Cambió estado a {'Activo' if status_update.is_active else 'Inactivo'} para el ejecutivo: {person_name}"
    )
    
    return {"detail": "Estado del ejecutivo actualizado."}


@router.delete("/{cred_id}")
def delete_credential(
    cred_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Elimina una credencial de la bóveda."""
    cred = db.query(Credential).filter(Credential.id == cred_id).first()
    if not cred:
        raise HTTPException(status_code=404, detail="Credencial no encontrada.")
        
    if not credential_is_visible_to(cred, current_user):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Acceso denegado.")

    title = cred.title
    db.delete(cred)
    db.commit()

    # Auditoría
    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="credential_delete",
        ip_address=request.client.host if request.client else None,
        details=f"Eliminó permanentemente la credencial '{title}' de la bóveda."
    )

    return {"detail": "Credencial eliminada exitosamente."}


@router.get("/template/download")
def download_credential_template(
    current_user: User = Depends(get_current_active_user)
):
    """Descarga una plantilla Excel vacía para la importación masiva de credenciales."""
    import openpyxl
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Credenciales"
    
    headers = ["Titulo", "URL", "Usuario", "Contraseña", "Categoria"]
    ws.append(headers)

    # Estilos básicos
    from openpyxl.styles import Font, PatternFill
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="049DD9", end_color="049DD9", fill_type="solid")
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25

    # Fila de ejemplo
    ws.append(["Servidor Producción", "https://prod.empresa.local", "admin", "P@ssw0rd123", "Servidores"])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {
        'Content-Disposition': 'attachment; filename="plantilla_credenciales.xlsx"'
    }
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)


@router.delete("/import/last")
def delete_last_import(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Elimina el último lote de credenciales importadas (agrupadas por timestamp de creación)."""
    from datetime import timedelta
    
    # Buscar la credencial más reciente del usuario
    latest = (
        db.query(Credential)
        .filter(Credential.owner_id == current_user.id)
        .order_by(Credential.created_at.desc())
        .first()
    )
    
    if not latest:
        raise HTTPException(status_code=404, detail="No hay importaciones previas para eliminar.")
    
    # Eliminar todas las credenciales creadas en el mismo minuto (± 60 segundos)
    window_start = latest.created_at - timedelta(seconds=60)
    window_end   = latest.created_at + timedelta(seconds=60)
    
    to_delete = (
        db.query(Credential)
        .filter(
            Credential.owner_id == current_user.id,
            Credential.created_at >= window_start,
            Credential.created_at <= window_end
        )
        .all()
    )
    
    count = len(to_delete)
    for cred in to_delete:
        db.delete(cred)
    db.commit()

    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="credential_import_delete",
        ip_address=request.client.host if request.client else None,
        details=f"Eliminó la última importación: {count} credenciales."
    )
    
    return {"detail": f"✅ Se eliminaron {count} credenciales del último lote importado.", "deleted": count}


@router.post("/import", status_code=status.HTTP_201_CREATED)
def import_credentials(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user)
):
    """Importa credenciales masivamente desde un archivo Excel con múltiples sistemas."""
    from app.services.credential_import_service import CredentialImportService

    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="El archivo debe ser un Excel (.xlsx)")

    try:
        contents = file.file.read()
        result = CredentialImportService.import_credentials_from_excel(
            file_content=contents,
            user_id=current_user.id,
            db=db
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail="Error procesando el archivo Excel.")

    # Log de auditoría
    if result["total_imported"] > 0:
        systems_str = ", ".join([f"{k}: {v}" for k, v in result["systems_imported"].items()])
        audit_service.log_action(
            db=db,
            user_id=current_user.id,
            action="credential_import",
            ip_address=request.client.host if request.client else None,
            details=f"Importación de ejecutivos: {result['total_imported']} credenciales. Sistemas: {systems_str}"
        )

    if result["total_errors"] > 0:
        return {
            "detail": f"Se importaron {result['total_imported']} credenciales con {result['total_errors']} errores.",
            "imported": result["total_imported"],
            "errors": result["errors"],
            "systems": result["systems_imported"]
        }
    
    return {
        "detail": f"✅ Importación exitosa: {result['total_imported']} credenciales de {len(result['systems_imported'])} sistemas.",
        "imported": result["total_imported"],
        "systems": result["systems_imported"]
    }


@router.post("/executive", status_code=status.HTTP_201_CREATED)
def create_executive_credentials(
    request: Request,
    exec_data: ExecutiveCredentialCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(PermissionChecker("credentials:manage"))
):
    """Crea credenciales para múltiples sistemas de un ejecutivo en una sola transacción."""
    imported_count = 0
    systems_created = []

    # Helper function to add/update a single credential
    def add_system_cred(title: str, username: str, password_raw: str, category: str, url_val: Optional[str] = None):
        nonlocal imported_count
        if not username or not password_raw:
            return
        
        encrypted_pw = crypto_service.encrypt_password(password_raw)
        
        # Check if already exists
        existing = db.query(Credential).filter(
            Credential.title == title,
            Credential.username == username,
            Credential.owner_id == current_user.id
        ).first()
        
        if existing:
            existing.url = url_val
            existing.encrypted_password = encrypted_pw
            existing.category = category
        else:
            db_cred = Credential(
                title=title,
                url=url_val,
                username=username,
                encrypted_password=encrypted_pw,
                category=category,
                owner_id=current_user.id
            )
            db.add(db_cred)
        
        imported_count += 1
        systems_created.append(category)

    try:
        # 0. Correo Personal
        if exec_data.personal_user and exec_data.personal_pass:
            add_system_cred(
                title=f"Correo Personal - {exec_data.name.strip().upper()}",
                username=exec_data.personal_user.strip(),
                password_raw=exec_data.personal_pass.strip(),
                category="Correo Corporativo"
            )

        # 1. Correo Corporativo
        if exec_data.correo_user and exec_data.correo_pass:
            add_system_cred(
                title=f"Correo - {exec_data.name.strip().upper()}",
                username=exec_data.correo_user.strip(),
                password_raw=exec_data.correo_pass.strip(),
                category="Correo Corporativo"
            )

        # 2. CRM
        if exec_data.crm_user and exec_data.crm_pass:
            add_system_cred(
                title=f"CRM - {exec_data.name.strip().upper()}",
                username=exec_data.crm_user.strip(),
                password_raw=exec_data.crm_pass.strip(),
                category="CRM"
            )

        # 3. Vocalcom
        if exec_data.vocalcom_user and exec_data.vocalcom_pass:
            add_system_cred(
                title=f"Vocalcom - {exec_data.name.strip().upper()}",
                username=exec_data.vocalcom_user.strip(),
                password_raw=exec_data.vocalcom_pass.strip(),
                category="Telefonía",
                url_val=exec_data.vocalcom_estacion.strip() if (exec_data.vocalcom_estacion and exec_data.vocalcom_estacion.strip()) else None
            )

        # 4. PC
        if exec_data.pc_user and exec_data.pc_pass:
            add_system_cred(
                title=f"PC - {exec_data.name.strip().upper()} ({exec_data.pc_user.strip()})",
                username=exec_data.pc_user.strip(),
                password_raw=exec_data.pc_pass.strip(),
                category="Sistemas"
            )

        if imported_count == 0:
            raise HTTPException(status_code=400, detail="Debe ingresar credenciales para al menos un sistema.")

        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creando credenciales de ejecutivo: {str(e)}")

    # Log de auditoría
    systems_str = ", ".join(systems_created)
    audit_service.log_action(
        db=db,
        user_id=current_user.id,
        action="credential_executive_create",
        ip_address=request.client.host if request.client else None,
        details=f"Creó credenciales agrupadas para ejecutivo '{exec_data.name}' en sistemas: {systems_str}"
    )

    return {
        "detail": f"Se registraron exitosamente {imported_count} credenciales para {exec_data.name.strip().upper()}.",
        "imported": imported_count,
        "systems": systems_created
    }
