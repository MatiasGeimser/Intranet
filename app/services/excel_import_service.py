from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError
from app.models.it_asset import ITAsset
from app.models.vlan import VLAN
from typing import Dict, List, Tuple
import os
from datetime import datetime


class ExcelImportService:
    """Servicio para importar datos de Excel (Software, VLANs, etc)"""

    @staticmethod
    def import_excel(file_path: str, db: Session, user_id: int) -> Dict:
        """
        Importa datos de un archivo Excel con múltiples hojas.
        
        Estructura esperada:
        - Hoja 'SOFTWARE': Columnas [Nombre, Variantes, Versión, Vendor, Estado]
        - Hoja 'VLAN': Columnas [ID VLAN, Nombre, Descripción, Red, Gateway, Estado]
        
        Args:
            file_path: Ruta del archivo Excel
            db: Sesión de base de datos
            user_id: ID del usuario que realiza la importación
            
        Returns:
            Dict con resumen de la importación
        """
        
        try:
            workbook = load_workbook(file_path)
            results = {
                "software_imported": 0,
                "software_errors": 0,
                "vlan_imported": 0,
                "vlan_errors": 0,
                "messages": []
            }
            
            # Procesar hoja de SOFTWARE
            if 'SOFTWARE' in workbook.sheetnames:
                results = ExcelImportService._import_software_sheet(
                    workbook, db, user_id, results
                )
            
            # Procesar hoja de VLAN
            if 'VLAN' in workbook.sheetnames:
                results = ExcelImportService._import_vlan_sheet(
                    workbook, db, user_id, results
                )
            
            db.commit()
            workbook.close()
            return results
            
        except Exception as e:
            db.rollback()
            raise Exception(f"Error al procesar Excel: {str(e)}")
    
    @staticmethod
    def _import_software_sheet(workbook, db: Session, user_id: int, results: Dict) -> Dict:
        """Importa la hoja de SOFTWARE"""
        
        sheet = workbook['SOFTWARE']
        row_number = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=False):  # Saltar encabezado
            row_number += 1
            
            try:
                # Extraer valores con manejo de None
                nombre = row[0].value
                variantes = row[1].value
                version = row[2].value if len(row) > 2 else None
                vendor = row[3].value if len(row) > 3 else None
                estado = row[4].value if len(row) > 4 else "Activo"
                
                # Validar campos obligatorios
                if not nombre:
                    results["software_errors"] += 1
                    results["messages"].append(
                        f"Fila {row_number} (SOFTWARE): Nombre es obligatorio"
                    )
                    continue
                
                # Crear descripción con variantes
                description = f"Variantes: {variantes}" if variantes else None
                
                # Verificar si ya existe
                existing = db.query(ITAsset).filter(
                    ITAsset.name == nombre,
                    ITAsset.asset_type == "SOFTWARE"
                ).first()
                
                if existing:
                    # Actualizar
                    existing.version = version or existing.version
                    existing.vendor = vendor or existing.vendor
                    existing.description = description or existing.description
                    existing.status = estado or existing.status
                    existing.updated_at = datetime.utcnow()
                else:
                    # Crear nuevo
                    new_software = ITAsset(
                        name=nombre,
                        asset_type="SOFTWARE",
                        category="Software",
                        description=description,
                        version=version,
                        vendor=vendor,
                        status=estado or "Activo",
                        created_by_id=user_id
                    )
                    db.add(new_software)
                
                results["software_imported"] += 1
                
            except IntegrityError:
                db.rollback()
                results["software_errors"] += 1
                results["messages"].append(
                    f"Fila {row_number} (SOFTWARE): Error de integridad en base de datos"
                )
            except Exception as e:
                results["software_errors"] += 1
                results["messages"].append(
                    f"Fila {row_number} (SOFTWARE): {str(e)}"
                )
        
        return results
    
    @staticmethod
    def _import_vlan_sheet(workbook, db: Session, user_id: int, results: Dict) -> Dict:
        """Importa la hoja de VLAN"""
        
        sheet = workbook['VLAN']
        row_number = 0
        
        for row in sheet.iter_rows(min_row=2, values_only=False):  # Saltar encabezado
            row_number += 1
            
            try:
                # Extraer valores
                vlan_id_val = row[0].value
                nombre = row[1].value
                descripcion = row[2].value if len(row) > 2 else None
                red = row[3].value if len(row) > 3 else None
                gateway = row[4].value if len(row) > 4 else None
                estado = row[5].value if len(row) > 5 else "Activo"
                
                # Validar campos obligatorios
                if not vlan_id_val or not nombre:
                    results["vlan_errors"] += 1
                    results["messages"].append(
                        f"Fila {row_number} (VLAN): ID VLAN y Nombre son obligatorios"
                    )
                    continue
                
                # Convertir a entero
                try:
                    vlan_id_val = int(vlan_id_val)
                except (ValueError, TypeError):
                    results["vlan_errors"] += 1
                    results["messages"].append(
                        f"Fila {row_number} (VLAN): ID VLAN debe ser numérico"
                    )
                    continue
                
                # Validar rango VLAN
                if not (1 <= vlan_id_val <= 4094):
                    results["vlan_errors"] += 1
                    results["messages"].append(
                        f"Fila {row_number} (VLAN): ID VLAN debe estar entre 1 y 4094"
                    )
                    continue
                
                # Verificar si ya existe
                existing = db.query(VLAN).filter(VLAN.vlan_id == vlan_id_val).first()
                
                if existing:
                    # Actualizar
                    existing.name = nombre
                    existing.description = descripcion or existing.description
                    existing.network = red or existing.network
                    existing.gateway = gateway or existing.gateway
                    existing.status = estado or existing.status
                    existing.updated_at = datetime.utcnow()
                else:
                    # Crear nueva
                    new_vlan = VLAN(
                        vlan_id=vlan_id_val,
                        name=nombre,
                        description=descripcion,
                        network=red,
                        gateway=gateway,
                        status=estado or "Activo",
                        created_by_id=user_id
                    )
                    db.add(new_vlan)
                
                results["vlan_imported"] += 1
                
            except IntegrityError:
                db.rollback()
                results["vlan_errors"] += 1
                results["messages"].append(
                    f"Fila {row_number} (VLAN): Error de integridad en base de datos"
                )
            except Exception as e:
                results["vlan_errors"] += 1
                results["messages"].append(
                    f"Fila {row_number} (VLAN): {str(e)}"
                )
        
        return results


excel_import_service = ExcelImportService()
