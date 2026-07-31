import csv
import io
import json
import os
import re
from sqlalchemy.orm import Session
from typing import Optional, List
from fastapi import UploadFile
from app.models.document import Document
from app.core.config import settings

try:
    import openpyxl
except ImportError:
    openpyxl = None

class DocService:
    @staticmethod
    def sanitize_filename(filename: str) -> str:
        """Sanitiza el nombre del archivo para evitar inyecciones de ruta o caracteres extraños."""
        # Remover partes de ruta
        filename = os.path.basename(filename)
        # Reemplazar espacios y caracteres inválidos
        filename = re.sub(r"[^\w\.-]", "_", filename)
        return filename

    @staticmethod
    def save_document(
        db: Session,
        upload_file: UploadFile,
        folder: str,
        uploader_id: int,
        is_public: bool = True,
        allowed_users_ids: List[int] = None
    ) -> Document:
        """Sube y guarda un archivo con versionado automático si ya existe en la misma carpeta."""
        original_name = DocService.sanitize_filename(upload_file.filename)
        base_name, ext = os.path.splitext(original_name)
        
        # Carpeta física en disco
        target_folder_path = os.path.join(settings.UPLOAD_DIR, folder)
        os.makedirs(target_folder_path, exist_ok=True)
        
        # Buscar coincidencias del mismo nombre en la misma carpeta virtual para el versionado
        existing_docs = db.query(Document).filter(
            Document.name == original_name,
            Document.folder == folder
        ).order_by(Document.created_at.desc()).all()
        
        version_num = 1
        if existing_docs:
            # Incrementar la versión
            latest_doc = existing_docs[0]
            try:
                # Extraer número de versión ej: "v2.0" -> 2
                match = re.match(r"v(\d+)\.0", latest_doc.version)
                if match:
                    version_num = int(match.group(1)) + 1
            except Exception:
                version_num = len(existing_docs) + 1
                
        version_str = f"v{version_num}.0"
        
        # Si la versión es > 1, modificamos el nombre del archivo guardado en disco
        saved_filename = original_name if version_num == 1 else f"{base_name}_{version_str}{ext}"
        file_path = os.path.join(target_folder_path, saved_filename)
        
        # Leer y escribir el archivo
        file_content = upload_file.file.read()
        size_bytes = len(file_content)
        
        with open(file_path, "wb") as f:
            f.write(file_content)
            
        # Crear entrada en BD
        # Guardar la ruta relativa para poder servirla por estáticos
        relative_path = os.path.relpath(file_path, start=".")
        
        db_doc = Document(
            name=original_name,
            file_path=relative_path.replace("\\", "/"),
            file_type=ext.lstrip(".").lower(),
            size_bytes=size_bytes,
            folder=folder,
            uploader_id=uploader_id,
            version=version_str,
            is_public=is_public
        )
        
        if not is_public and allowed_users_ids:
            from app.models.user import User
            users = db.query(User).filter(User.id.in_(allowed_users_ids)).all()
            db_doc.allowed_users.extend(users)
            
        db.add(db_doc)
        db.commit()
        db.refresh(db_doc)
        
        return db_doc

    @staticmethod
    def save_storage_document(
        db: Session,
        object_path: str,
        original_name: str,
        folder: str,
        uploader_id: int,
        size_bytes: int,
        file_type: str,
        is_public: bool = True,
        allowed_users_ids: List[int] = None,
    ) -> Document:
        """Registra un objeto ya cargado directamente en Supabase Storage."""
        original_name = DocService.sanitize_filename(original_name)[:100]
        existing_docs = db.query(Document).filter(
            Document.name == original_name,
            Document.folder == folder,
        ).order_by(Document.created_at.desc()).all()
        version_num = 1
        if existing_docs:
            match = re.match(r"v(\d+)\.0", existing_docs[0].version or "")
            version_num = int(match.group(1)) + 1 if match else len(existing_docs) + 1

        db_doc = Document(
            name=original_name,
            file_path=object_path,
            file_type=file_type,
            size_bytes=size_bytes,
            folder=folder,
            uploader_id=uploader_id,
            version=f"v{version_num}.0",
            is_public=is_public,
        )
        if not is_public and allowed_users_ids:
            from app.models.user import User
            users = db.query(User).filter(User.id.in_(allowed_users_ids)).all()
            db_doc.allowed_users.extend(users)
        db.add(db_doc)
        db.commit()
        db.refresh(db_doc)
        return db_doc

    @staticmethod
    def read_document_preview(doc: Document) -> dict:
        """Genera los datos necesarios para visualizar o editar un documento."""
        file_path = os.path.normpath(doc.file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError("El archivo físico no existe en el servidor.")

        text_types = {"txt", "md", "json", "key", "pem", "pfx", "crt", "cer", "csv"}
        binary_preview = {"pdf", "png", "jpg", "jpeg", "gif", "webp"}

        if doc.file_type in text_types:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()
            return {
                "mode": "text",
                "file_type": doc.file_type,
                "content": content
            }

        if doc.file_type == "xlsx":
            if openpyxl is None:
                raise RuntimeError("La vista previa de archivos XLSX requiere la dependencia openpyxl.")

            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets = []
            for worksheet in workbook.worksheets:
                rows = []
                for row in worksheet.iter_rows(values_only=True):
                    rows.append(["" if cell is None else str(cell) for cell in row])
                sheets.append({
                    "name": worksheet.title,
                    "rows": rows
                })
            return {
                "mode": "table",
                "file_type": doc.file_type,
                "sheets": sheets
            }

        if doc.file_type in binary_preview:
            return {
                "mode": "binary",
                "file_type": doc.file_type,
                "preview_url": "/" + doc.file_path.replace('\\\\', '/')
            }

        return {
            "mode": "binary",
            "file_type": doc.file_type,
            "preview_url": "/" + doc.file_path.replace('\\\\', '/')
        }

    @staticmethod
    def save_document_content(db: Session, doc: Document, content: Optional[str] = None, rows: Optional[List[List[str]]] = None, sheet_name: Optional[str] = None) -> Document:
        """Guarda cambios directos sobre el documento editado por el administrador."""
        file_path = os.path.normpath(doc.file_path)
        if not os.path.exists(file_path):
            raise FileNotFoundError("El archivo físico no existe en el servidor.")

        if doc.file_type in {"txt", "md", "json", "key", "pem", "pfx", "crt", "cer", "csv"}:
            if content is None:
                raise ValueError("No se recibió contenido para guardar.")
            with open(file_path, "w", encoding="utf-8", newline="") as f:
                f.write(content)
        elif doc.file_type == "xlsx":
            if openpyxl is None:
                raise RuntimeError("La edición de archivos XLSX requiere la dependencia openpyxl.")
            if rows is None:
                raise ValueError("No se recibieron filas para guardar el XLSX.")

            workbook = openpyxl.load_workbook(file_path)
            target_sheet = None
            if sheet_name and sheet_name in workbook.sheetnames:
                target_sheet = workbook[sheet_name]
            elif sheet_name:
                target_sheet = workbook.create_sheet(sheet_name)
            else:
                target_sheet = workbook.active

            if target_sheet.max_row > 0:
                target_sheet.delete_rows(1, target_sheet.max_row)

            for row in rows:
                target_sheet.append(row)
            workbook.save(file_path)
        else:
            raise ValueError("No es posible editar este tipo de archivo directamente desde la vista.")

        doc.size_bytes = os.path.getsize(file_path)
        db.commit()
        db.refresh(doc)
        return doc

    @staticmethod
    def get_documents_by_folder(db: Session, folder: str) -> List[Document]:
        """Obtiene la lista de documentos en una carpeta virtual."""
        return db.query(Document).filter(Document.folder == folder).order_by(Document.created_at.desc()).all()

doc_service = DocService()
