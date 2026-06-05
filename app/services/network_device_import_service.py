from openpyxl import load_workbook
from sqlalchemy.orm import Session
from app.models.network_devices import SwitchDevice, SwitchInterface
from app.models.vlan import VLAN
from typing import Dict, List, Tuple, Optional
import re


class NetworkDeviceImportService:
    """Servicio para importar configuración de redes desde Excel de Cisco"""

    @staticmethod
    def import_switch_config(file_path: str, db: Session, user_id: int) -> Dict:
        """
        Importa configuración completa de un Switch desde Excel.
        
        Formato esperado:
        - Fila 1-2: Info del switch (HOSTNAME, IP)
        - Fila 4: Encabezados
        - Fila 5+: Datos de VLANs (izq) e Interfaces (der)
        """
        
        try:
            workbook = load_workbook(file_path)
            
            results = {
                "switch_created": False,
                "switch_id": None,
                "switch_ids": [],
                "vlans_imported": 0,
                "vlans_errors": 0,
                "interfaces_imported": 0,
                "interfaces_errors": 0,
                "messages": []
            }
            
            # Procesar todas las hojas del workbook
            for ws in workbook.worksheets:
                results["messages"].append(f"Procesando hoja: {ws.title}")
                try:
                    hostname = ws.cell(2, 1).value or "UNKNOWN"
                    ip_address = ws.cell(2, 2).value or "0.0.0.0"
                    model = ws.title if ws.title else "Cisco Switch"
                    
                    switch = db.query(SwitchDevice).filter(
                        SwitchDevice.hostname == hostname
                    ).first()
                    
                    if not switch:
                        switch = SwitchDevice(
                            hostname=hostname,
                            ip_address=str(ip_address),
                            model=model,
                            created_by_id=user_id
                        )
                        db.add(switch)
                        db.flush()
                        results["switch_created"] = True
                        results["messages"].append(f"✓ Switch creado: {hostname}")
                    else:
                        switch.ip_address = str(ip_address)
                        results["messages"].append(f"✓ Switch actualizado: {hostname}")
                    
                    results["switch_id"] = switch.id
                    results["switch_ids"].append(switch.id)
                    
                    # Extraer VLANs (columnas A-C, desde fila 5)
                    # y Interfaces (columnas E-G, desde fila 5)
                    for row_idx in range(5, ws.max_row + 1):
                        # Procesar VLANs (izquierda)
                        vlan_id_cell = ws.cell(row_idx, 1).value
                        vlan_name = ws.cell(row_idx, 2).value
                        vlan_status = ws.cell(row_idx, 3).value
                        
                        if vlan_id_cell and vlan_name:
                            try:
                                vlan_id_num = int(vlan_id_cell) if isinstance(vlan_id_cell, int) else int(str(vlan_id_cell).split()[0])
                                
                                # Buscar o crear VLAN
                                existing_vlan = db.query(VLAN).filter(VLAN.vlan_id == vlan_id_num).first()
                                
                                if not existing_vlan:
                                    new_vlan = VLAN(
                                        vlan_id=vlan_id_num,
                                        name=str(vlan_name),
                                        status="Activo" if vlan_status and "active" in str(vlan_status).lower() else "Inactivo",
                                        created_by_id=user_id
                                    )
                                    db.add(new_vlan)
                                    db.flush()
                                    results["vlans_imported"] += 1
                                else:
                                    existing_vlan.name = str(vlan_name)
                                    existing_vlan.status = "Activo" if vlan_status and "active" in str(vlan_status).lower() else "Inactivo"
                                    results["vlans_imported"] += 1
                            
                            except (ValueError, TypeError) as e:
                                results["vlans_errors"] += 1
                                results["messages"].append(f"⚠ Fila {row_idx}: Error VLAN - {str(e)}")
                        
                        # Procesar Interfaces (derecha: columnas E-G)
                        interface_name = ws.cell(row_idx, 5).value
                        vlan_port = ws.cell(row_idx, 6).value
                        description = ws.cell(row_idx, 7).value
                        
                        if interface_name:
                            try:
                                interface_name_str = str(interface_name).strip()
                                
                                # Detectar tipo de puerto
                                port_type = "Gigabit" if "Gi" in interface_name_str else "FastEthernet"
                                is_uplink = "Gi" in interface_name_str
                                is_trunk = vlan_port and "Trunk" in str(vlan_port)
                                
                                # Extraer número de puerto como entero (parte derecha de Fa0/1, Gi1/0, etc.)
                                port_match = re.search(r'(?:\D*)(\d+)[/\.](\d+)$', interface_name_str)
                                port_number = int(port_match.group(2)) if port_match else None
                                
                                # Detectar tipo de dispositivo desde descripción
                                device_type = NetworkDeviceImportService._detect_device_type(description)
                                
                                # Buscar interfaz existente
                                existing_interface = db.query(SwitchInterface).filter(
                                    SwitchInterface.switch_id == switch.id,
                                    SwitchInterface.interface_name == interface_name_str
                                ).first()
                                
                                if not existing_interface:
                                    interface = SwitchInterface(
                                        switch_id=switch.id,
                                        interface_name=interface_name_str,
                                        port_type=port_type,
                                        port_number=port_number,
                                        vlan_name=str(vlan_port) if vlan_port else None,
                                        description=str(description) if description else None,
                                        is_uplink=is_uplink,
                                        is_trunk=is_trunk,
                                        status="Active",
                                        connected_device_type=device_type,
                                        created_by_id=user_id
                                    )
                                    db.add(interface)
                                    results["interfaces_imported"] += 1
                                else:
                                    existing_interface.vlan_name = str(vlan_port) if vlan_port else None
                                    existing_interface.description = str(description) if description else None
                                    existing_interface.port_type = port_type
                                    existing_interface.is_uplink = is_uplink
                                    existing_interface.is_trunk = is_trunk
                                    existing_interface.connected_device_type = device_type
                                    results["interfaces_imported"] += 1
                            
                            except Exception as e:
                                results["interfaces_errors"] += 1
                                results["messages"].append(f"⚠ Hoja {ws.title} fila {row_idx}: Error Interface '{interface_name}' - {str(e)}")
                except Exception as e:
                    results["messages"].append(f"⚠ Hoja {ws.title}: {str(e)}")
            
            db.commit()
            workbook.close()
            
            return results
        
        except Exception as e:
            db.rollback()
            raise Exception(f"Error al procesar configuración de red: {str(e)}")
    
    @staticmethod
    def _detect_device_type(description: Optional[str]) -> Optional[str]:
        """Detecta tipo de dispositivo desde la descripción"""
        if not description:
            return None
        
        desc_lower = str(description).lower()
        
        type_patterns = {
            "Switch": ["sw-", "switch"],
            "Router": ["router"],
            "Access Point": ["ap", "wifi", "wireless"],
            "Camera": ["camera", "cam"],
            "Printer": ["printer", "print"],
            "Phone": ["phone", "voip"],
            "Server": ["server", "svr"],
            "PC": ["endpoint", "pc", "desktop"],
            "Unknown": []
        }
        
        for device_type, patterns in type_patterns.items():
            if any(pattern in desc_lower for pattern in patterns):
                return device_type
        
        return "Unknown"


network_device_import_service = NetworkDeviceImportService()
