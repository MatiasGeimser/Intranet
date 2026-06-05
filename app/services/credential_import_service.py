"""
Servicio especializado para importación de credenciales desde Excel.
Soporta el formato de credenciales de ejecutivos con múltiples sistemas.
"""
import openpyxl
import io
from typing import Dict, List, Any
from sqlalchemy.orm import Session
from app.models.credential import Credential
from app.services.crypto_service import crypto_service


class CredentialImportService:
    """Importa credenciales desde Excel con múltiples formatos."""

    @staticmethod
    def import_credentials_from_excel(
        file_content: bytes,
        user_id: int,
        db: Session
    ) -> Dict[str, Any]:
        """
        Importa credenciales desde Excel.
        Soporta formato de ejecutivos con múltiples sistemas por persona.

        Columnas esperadas:
        Nombres | Correo Info y Usuario CRM | Clave Info | Clave CRM |
        Usuario Vocalcom | Password Vocalcom | Estacion Vocalcom |
        Usuario PC | Clave PC | Estado
        """
        try:
            wb = openpyxl.load_workbook(filename=io.BytesIO(file_content), data_only=True)
        except Exception as e:
            raise ValueError(f"Error al leer archivo Excel: {str(e)}")

        results = {
            "total_imported": 0,
            "total_errors": 0,
            "systems_imported": {},
            "errors": []
        }

        # Procesar todas las hojas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                continue

            # Buscar la fila de cabecera real
            header_row_idx = 0
            headers = []
            for i, row in enumerate(rows[:5]):
                row_vals = [str(cell).strip().lower() if cell is not None else "" for cell in row]
                if any(h in ["nombres", "nombre", "titulo", "title"] for h in row_vals) or \
                   any("vocalcom" in h for h in row_vals):
                    header_row_idx = i
                    headers = row_vals
                    break

            if not headers:
                headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
                header_row_idx = 0

            # Intentar mapear las columnas estándar
            col_map = {}
            for idx, h in enumerate(headers):
                if h in ["titulo", "title", "nombre", "nombre del servicio"]:
                    col_map["title"] = idx
                elif h in ["url", "link", "enlace", "direccion", "dirección"]:
                    col_map["url"] = idx
                elif h in ["usuario", "username", "login", "user"]:
                    col_map["username"] = idx
                elif h in ["contraseña", "contrasena", "password", "clave", "pass", "pw"]:
                    col_map["password"] = idx
                elif h in ["categoria", "categoría", "category", "grupo"]:
                    col_map["category"] = idx

            # Intentar mapear formato simplificado de Vocalcom
            vocalcom_map = {}
            for idx, h in enumerate(headers):
                if h in ["nombres", "nombre", "persona", "ejecutivo", "title", "titulo"]:
                    vocalcom_map["name"] = idx
                elif "usuario, password y estacion vocalcom" in h or h == "vocalcom":
                    vocalcom_map["vocalcom"] = idx
                elif h in ["usuario", "usuario vocalcom", "user", "username"]:
                    vocalcom_map["username"] = idx
                elif h in ["contraseña", "contrasena", "password", "clave", "pass", "pw", "clave vocalcom"]:
                    vocalcom_map["password"] = idx
                elif h in ["estacion", "estación", "estacion vocalcom", "vocalcom estacion",
                           "estacion de vocalcom", "extensión", "extension"]:
                    vocalcom_map["estacion"] = idx

            is_vocalcom_simplified = (
                "name" in vocalcom_map and
                ("vocalcom" in vocalcom_map or
                 ("username" in vocalcom_map and "password" in vocalcom_map and "estacion" in vocalcom_map)) and
                not any(x in headers for x in ["correo", "pc", "clave pc",
                                                "correo info y usuario crm", "correo info"])
            )

            is_standard = ("title" in col_map and "username" in col_map and "password" in col_map)

            if is_vocalcom_simplified:
                # Procesar formato simplificado de Vocalcom
                for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
                    name_idx = vocalcom_map["name"]
                    if len(row) <= name_idx or row[name_idx] is None or not str(row[name_idx]).strip():
                        continue

                    try:
                        person_name = str(row[name_idx]).strip()

                        if "vocalcom" in vocalcom_map:
                            vocalcom_col = vocalcom_map["vocalcom"]
                            username = str(row[vocalcom_col]).strip() if (len(row) > vocalcom_col and row[vocalcom_col] is not None) else ""
                            password = str(row[vocalcom_col + 1]).strip() if (len(row) > vocalcom_col + 1 and row[vocalcom_col + 1] is not None) else ""
                            estacion = str(row[vocalcom_col + 2]).strip() if (len(row) > vocalcom_col + 2 and row[vocalcom_col + 2] is not None) else ""
                        else:
                            username_idx = vocalcom_map["username"]
                            username = str(row[username_idx]).strip() if (len(row) > username_idx and row[username_idx] is not None) else ""

                            password_idx = vocalcom_map["password"]
                            password = str(row[password_idx]).strip() if (len(row) > password_idx and row[password_idx] is not None) else ""

                            estacion_idx = vocalcom_map["estacion"]
                            estacion = str(row[estacion_idx]).strip() if (len(row) > estacion_idx and row[estacion_idx] is not None) else ""

                        if username and password:
                            encrypted_pw = crypto_service.encrypt_password(password)
                            title = f"Vocalcom - {person_name}"
                            url_val = estacion if (estacion and str(estacion).lower() != 'none') else None

                            existing = db.query(Credential).filter(
                                Credential.title == title,
                                Credential.username == username,
                                Credential.owner_id == user_id
                            ).first()

                            if existing:
                                existing.url = url_val
                                existing.encrypted_password = encrypted_pw
                                existing.category = "Telefonía"
                            else:
                                db_cred = Credential(
                                    title=title,
                                    url=url_val,
                                    username=username,
                                    encrypted_password=encrypted_pw,
                                    category="Telefonía",
                                    owner_id=user_id
                                )
                                db.add(db_cred)

                            results["systems_imported"]["Telefonía"] = results["systems_imported"].get("Telefonía", 0) + 1
                            results["total_imported"] += 1
                    except Exception as e:
                        results["total_errors"] += 1
                        results["errors"].append(f"Fila {row_idx}: {str(e)}")

            elif is_standard:
                # Procesar formato estándar
                for row_idx, row in enumerate(rows[1:], start=2):
                    title_idx = col_map["title"]
                    if len(row) <= title_idx or row[title_idx] is None or not str(row[title_idx]).strip():
                        continue

                    try:
                        title = str(row[title_idx]).strip()

                        username_idx = col_map["username"]
                        username = str(row[username_idx]).strip() if (len(row) > username_idx and row[username_idx] is not None) else ""

                        password_idx = col_map["password"]
                        password = str(row[password_idx]).strip() if (len(row) > password_idx and row[password_idx] is not None) else ""

                        url_idx = col_map.get("url")
                        url = str(row[url_idx]).strip() if (url_idx is not None and len(row) > url_idx and row[url_idx] is not None) else None

                        category_idx = col_map.get("category")
                        category = str(row[category_idx]).strip() if (category_idx is not None and len(row) > category_idx and row[category_idx] is not None) else "General"

                        if username and password:
                            encrypted_pw = crypto_service.encrypt_password(password)

                            existing = db.query(Credential).filter(
                                Credential.title == title,
                                Credential.username == username,
                                Credential.owner_id == user_id
                            ).first()

                            if existing:
                                existing.url = url
                                existing.encrypted_password = encrypted_pw
                                existing.category = category
                            else:
                                db_cred = Credential(
                                    title=title,
                                    url=url,
                                    username=username,
                                    encrypted_password=encrypted_pw,
                                    category=category,
                                    owner_id=user_id
                                )
                                db.add(db_cred)

                            results["systems_imported"][category] = results["systems_imported"].get(category, 0) + 1
                            results["total_imported"] += 1
                    except Exception as e:
                        results["total_errors"] += 1
                        results["errors"].append(f"Fila {row_idx}: {str(e)}")

            else:
                # Procesar cada fila (ejecutivo) - Formato multi-sistema
                # Calcular mapeo de índices dinámicamente desde las cabeceras
                idx_map = {
                    "correo_crm": 1,
                    "clave_info": 2,
                    "clave_crm": 3,
                    "vocalcom_user": 4,
                    "vocalcom_pass": 5,
                    "vocalcom_estacion": 6,
                    "pc_user": 7,
                    "pc_pass": 8,
                    "estado": 9
                }

                for idx, h in enumerate(headers):
                    if not h:
                        continue
                    h_clean = str(h).strip().lower()
                    if "correo info" in h_clean:
                        idx_map["correo_crm"] = idx
                    elif "clave info" in h_clean:
                        idx_map["clave_info"] = idx
                    elif "clave crm" in h_clean:
                        idx_map["clave_crm"] = idx
                    elif "usuario vocalcom" in h_clean or "usuario, password y estacion vocalcom" in h_clean:
                        idx_map["vocalcom_user"] = idx
                        idx_map["vocalcom_pass"] = idx + 1
                        idx_map["vocalcom_estacion"] = idx + 2
                    elif "password vocalcom" in h_clean or "clave vocalcom" in h_clean:
                        idx_map["vocalcom_pass"] = idx
                    elif "estacion vocalcom" in h_clean or h_clean in ["estacion", "estación"]:
                        idx_map["vocalcom_estacion"] = idx
                    elif "usuario pc" in h_clean:
                        idx_map["pc_user"] = idx
                    elif "clave pc" in h_clean:
                        idx_map["pc_pass"] = idx
                    elif h_clean in ["estado", "status"]:
                        idx_map["estado"] = idx

                for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
                    if not row[0]:  # Si no hay nombre, saltar
                        continue

                    # Saltar desvinculados
                    estado_idx = idx_map.get("estado", 9)
                    if len(row) > estado_idx and row[estado_idx] == "Desvinculado":
                        continue

                    try:
                        person_name = str(row[0]).strip()

                        credentials_to_import = CredentialImportService._extract_credentials(
                            person_name, row, idx_map
                        )

                        for cred in credentials_to_import:
                            if cred["username"] and cred["password"]:
                                encrypted_pw = crypto_service.encrypt_password(cred["password"])

                                existing = db.query(Credential).filter(
                                    Credential.title == cred["title"],
                                    Credential.username == cred["username"],
                                    Credential.owner_id == user_id
                                ).first()

                                if existing:
                                    existing.url = cred.get("url")
                                    existing.encrypted_password = encrypted_pw
                                    existing.category = cred["category"]
                                else:
                                    db_cred = Credential(
                                        title=cred["title"],
                                        url=cred.get("url"),
                                        username=cred["username"],
                                        encrypted_password=encrypted_pw,
                                        category=cred["category"],
                                        owner_id=user_id
                                    )
                                    db.add(db_cred)

                                system = cred["category"]
                                results["systems_imported"][system] = results["systems_imported"].get(system, 0) + 1
                                results["total_imported"] += 1

                    except Exception as e:
                        results["total_errors"] += 1
                        results["errors"].append(f"Fila {row_idx}: {str(e)}")

        if results["total_imported"] > 0:
            db.commit()

        return results

    @staticmethod
    def _extract_credentials(person_name: str, row: tuple, idx_map: dict = None) -> List[Dict[str, str]]:
        """
        Extrae credenciales del formato de ejecutivos.
        Mapea los datos a credenciales de diferentes sistemas.
        """
        credentials = []

        if idx_map is None:
            idx_map = {
                "correo_crm": 1,
                "clave_info": 2,
                "clave_crm": 3,
                "vocalcom_user": 4,
                "vocalcom_pass": 5,
                "vocalcom_estacion": 6,
                "pc_user": 7,
                "pc_pass": 8,
                "estado": 9
            }

        try:
            # Sistema 1: Correo Corporativo
            c_idx = idx_map.get("correo_crm", 1)
            p_idx = idx_map.get("clave_info", 2)
            if len(row) > c_idx and row[c_idx]:
                correo = str(row[c_idx]).strip()
                if len(row) > p_idx and row[p_idx]:
                    correo_pass = str(row[p_idx]).strip()
                    if correo and correo_pass and correo_pass.lower() != 'none':
                        credentials.append({
                            "title": f"Correo - {person_name}",
                            "username": correo,
                            "password": correo_pass,
                            "category": "Correo Corporativo",
                            "url": None
                        })

            # Sistema 2: CRM
            c_idx = idx_map.get("correo_crm", 1)
            p_idx = idx_map.get("clave_crm", 3)
            if len(row) > c_idx and row[c_idx]:
                crm_user = str(row[c_idx]).strip()
                if len(row) > p_idx and row[p_idx]:
                    crm_pass = str(row[p_idx]).strip()
                    if crm_user and crm_pass and crm_pass.lower() != 'none':
                        credentials.append({
                            "title": f"CRM - {person_name}",
                            "username": crm_user,
                            "password": crm_pass,
                            "category": "CRM",
                            "url": None
                        })

            # Sistema 3: Vocalcom (Telefonía)
            u_idx = idx_map.get("vocalcom_user", 4)
            p_idx = idx_map.get("vocalcom_pass", 5)
            e_idx = idx_map.get("vocalcom_estacion", 6)
            if len(row) > u_idx and row[u_idx]:
                vocalcom_user = str(row[u_idx]).strip()
                if len(row) > p_idx and row[p_idx]:
                    vocalcom_pass = str(row[p_idx]).strip()
                    estacion = (
                        str(row[e_idx]).strip()
                        if (len(row) > e_idx and row[e_idx] is not None and
                            str(row[e_idx]).strip() and str(row[e_idx]).strip().lower() != 'none')
                        else None
                    )
                    if vocalcom_user and vocalcom_pass and vocalcom_pass.lower() != 'none':
                        credentials.append({
                            "title": f"Vocalcom - {person_name}",
                            "username": vocalcom_user,
                            "password": vocalcom_pass,
                            "category": "Telefonía",
                            "url": estacion
                        })

            # Sistema 4: PC (Sistemas)
            u_idx = idx_map.get("pc_user", 7)
            p_idx = idx_map.get("pc_pass", 8)
            if len(row) > u_idx and row[u_idx]:
                pc_name = str(row[u_idx]).strip()
                if len(row) > p_idx and row[p_idx]:
                    pc_pass = str(row[p_idx]).strip()
                    if pc_name and pc_pass and pc_pass.lower() not in ('none', 'tiene'):
                        credentials.append({
                            "title": f"PC - {person_name} ({pc_name})",
                            "username": pc_name,
                            "password": pc_pass,
                            "category": "Sistemas",
                            "url": None
                        })

        except Exception as e:
            print(f"Error extrayendo credenciales para {person_name}: {e}")

        return credentials
