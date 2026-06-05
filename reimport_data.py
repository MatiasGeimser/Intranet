import sys
import os
import openpyxl

sys.path.append("c:\\Intranet")

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from app.core.config import settings
from app.models.credential import Credential

# Initialize DB Session
engine = create_engine(settings.DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
db = SessionLocal()

excel_path = r"c:\Intranet\static\uploads\General\Usuarios Equifax Placa.xlsx"

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print("Sheets in workbook:", wb.sheetnames)
    ws = wb['Equifax']
    rows = list(ws.iter_rows(values_only=True))
    print(f"Total rows in Equifax sheet: {len(rows)}")
    print("Row 0:", rows[0])
    print("Row 1:", rows[1])
    print("Row 2:", rows[2])

    # Let's run a manual simulation of index mapping
    headers = []
    header_row_idx = 0
    for i, row in enumerate(rows[:5]):
        row_vals = [str(cell).strip().lower() if cell is not None else "" for cell in row]
        if any(h in ["nombres", "nombre", "titulo", "title"] for h in row_vals) or \
           any("vocalcom" in h for h in row_vals):
            header_row_idx = i
            headers = row_vals
            break
            
    print("Detected header_row_idx:", header_row_idx)
    print("Detected headers:", headers)

    idx_map = {
        "correo_crm": 1,
        "clave_info": 2,
        "clave_crm": 3,
        "vocalcom_user": 4,
        "vocalcom_pass": 5,
        "vocalcom_estacion": 6,
        "pc_user": 7,
        "pc_pass": 8,
        "estado": 9
    }
    
    for idx, h in enumerate(headers):
        if not h:
            continue
        h_clean = str(h).strip().lower()
        if "correo info" in h_clean:
            idx_map["correo_crm"] = idx
        elif "clave info" in h_clean:
            idx_map["clave_info"] = idx
        elif "clave crm" in h_clean:
            idx_map["clave_crm"] = idx
        elif "usuario vocalcom" in h_clean or "usuario, password y estacion vocalcom" in h_clean or (h_clean == "vocalcom" and idx_map.get("vocalcom_user") == 4):
            idx_map["vocalcom_user"] = idx
            idx_map["vocalcom_pass"] = idx + 1
            idx_map["vocalcom_estacion"] = idx + 2
        elif "password vocalcom" in h_clean or "clave vocalcom" in h_clean:
            idx_map["vocalcom_pass"] = idx
        elif "estacion vocalcom" in h_clean or h_clean in ["estacion", "estación"]:
            idx_map["vocalcom_estacion"] = idx
        elif "usuario pc" in h_clean or h_clean == "pc":
            idx_map["pc_user"] = idx
        elif "clave pc" in h_clean:
            idx_map["pc_pass"] = idx
        elif h_clean in ["estado", "status"]:
            idx_map["estado"] = idx

    print("Computed idx_map:", idx_map)

    # Let's see what is imported
    from app.services.credential_import_service import CredentialImportService
    for row_idx, row in enumerate(rows[header_row_idx + 1:], start=header_row_idx + 2):
        if not row[0]:
            continue
        estado_idx = idx_map.get("estado", 9)
        if len(row) > estado_idx and row[estado_idx] == "Desvinculado":
            print(f"Row {row_idx}: skipping desvinculado")
            continue
        person_name = str(row[0]).strip()
        creds = CredentialImportService._extract_credentials(person_name, row, idx_map)
        print(f"Row {row_idx} ({person_name}): extracted {len(creds)} credentials")
        for c in creds:
            print(f"  - {c['category']} | {c['title']} | {c['username']} | {c['password']}")

finally:
    db.close()
