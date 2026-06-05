#!/usr/bin/env python3
"""
Script para generar archivo Excel de ejemplo para importación de Software y VLANs
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def create_example_excel(output_file='ejemplo_importacion.xlsx'):
    """
    Crea un archivo Excel de ejemplo con dos hojas:
    - SOFTWARE: Datos de software con variantes
    - VLAN: Configuración de VLANs
    """
    
    wb = Workbook()
    ws_software = wb.active
    ws_software.title = "SOFTWARE"
    
    # Crear hoja VLAN
    ws_vlan = wb.create_sheet("VLAN")
    
    # ═══════════ HOJA SOFTWARE ═══════════
    
    # Encabezados
    headers_sw = ["Nombre", "Variantes", "Versión", "Vendor", "Estado"]
    ws_software.append(headers_sw)
    
    # Estilizar encabezado
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws_software[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos de ejemplo
    software_data = [
        ["Microsoft Office 365", "Pro, Standard, Home", "2024", "Microsoft", "Activo"],
        ["Adobe Creative Cloud", "Premium, Student", "2024", "Adobe", "Activo"],
        ["Visual Studio Code", "", "1.95", "Microsoft", "Activo"],
        ["Python", "3.11, 3.12", "3.12", "Python Software Foundation", "Activo"],
        ["PostgreSQL", "", "16", "PostgreSQL", "Activo"],
        ["Docker", "Desktop, Enterprise", "24.0", "Docker", "Activo"],
        ["Slack", "", "4.38", "Slack", "Activo"],
        ["Zoom", "Basic, Pro, Business", "5.16", "Zoom", "Activo"],
    ]
    
    for row in software_data:
        ws_software.append(row)
    
    # Ajustar ancho de columnas
    ws_software.column_dimensions['A'].width = 25
    ws_software.column_dimensions['B'].width = 30
    ws_software.column_dimensions['C'].width = 15
    ws_software.column_dimensions['D'].width = 25
    ws_software.column_dimensions['E'].width = 15
    
    # Alineación y bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws_software.iter_rows(min_row=1, max_row=ws_software.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # ═══════════ HOJA VLAN ═══════════
    
    # Encabezados
    headers_vlan = ["ID VLAN", "Nombre", "Descripción", "Red", "Gateway", "Estado"]
    ws_vlan.append(headers_vlan)
    
    for cell in ws_vlan[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Datos de ejemplo
    vlan_data = [
        [10, "Administración", "Red de administradores de sistemas", "192.168.1.0/24", "192.168.1.1", "Activo"],
        [20, "Usuarios", "Red general de usuarios de oficina", "192.168.2.0/24", "192.168.2.1", "Activo"],
        [30, "Invitados", "Red de acceso para visitantes", "192.168.3.0/24", "192.168.3.1", "Activo"],
        [40, "Servidores", "Red de infraestructura de servidores", "192.168.4.0/24", "192.168.4.1", "Activo"],
        [50, "IoT", "Red para dispositivos IoT", "192.168.5.0/24", "192.168.5.1", "Activo"],
        [100, "Desarrollo", "Red de desarrollo y pruebas", "10.0.1.0/24", "10.0.1.1", "Activo"],
        [101, "Producción", "Red de sistemas en producción", "10.0.2.0/24", "10.0.2.1", "Activo"],
    ]
    
    for row in vlan_data:
        ws_vlan.append(row)
    
    # Ajustar ancho de columnas
    ws_vlan.column_dimensions['A'].width = 12
    ws_vlan.column_dimensions['B'].width = 18
    ws_vlan.column_dimensions['C'].width = 35
    ws_vlan.column_dimensions['D'].width = 18
    ws_vlan.column_dimensions['E'].width = 18
    ws_vlan.column_dimensions['F'].width = 12
    
    # Alineación y bordes
    for row in ws_vlan.iter_rows(min_row=1, max_row=ws_vlan.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            if cell.row > 1:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Guardar
    wb.save(output_file)
    print(f"✓ Archivo de ejemplo creado: {output_file}")
    print(f"  - Hoja SOFTWARE: {len(software_data)} ejemplos")
    print(f"  - Hoja VLAN: {len(vlan_data)} ejemplos")


if __name__ == "__main__":
    create_example_excel()
